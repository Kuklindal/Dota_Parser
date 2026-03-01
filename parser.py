# -*- coding: utf-8 -*-
"""
Парсер статистики матчей Dota 2 с ru.dotabuff.com (Dotabuff + OpenDota fallback).
- Dotabuff: heroes, K/D/A, GPM/XPM/LH/DN, lanes tab.
- OpenDota: надёжный fallback для lane/lane_role/position и талантов (lvl10).
"""

import re
import time
import random
import logging
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook, Workbook

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

BASE_URL = "https://ru.dotabuff.com"
OPENDOTA_URL = "https://api.opendota.com/api"

USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
_playwright_storage_state = None


def fetch_dotabuff_with_playwright(url, timeout=60000, save_debug_path=None, headed=False, max_total_ms=120000, logger=None):
    """
    Надёжная загрузка страниц Dotabuff через Playwright.
    Важно: НЕ используем networkidle (на Dotabuff часто бесконечные запросы).
    """
    if not HAS_PLAYWRIGHT:
        raise RuntimeError("Нужен Playwright: pip install playwright && playwright install chromium")

    global _playwright_storage_state
    started = time.monotonic()

    def _remaining_ms():
        if max_total_ms is None:
            return None
        elapsed = int((time.monotonic() - started) * 1000)
        return max_total_ms - elapsed

    def _clamp_timeout(ms, min_ms=1000):
        rem = _remaining_ms()
        if rem is None:
            return ms
        if rem <= 0:
            raise TimeoutError(f"Timeout while loading {url}")
        if rem < min_ms:
            return rem
        return min(ms, rem)

    stealth_script = """
    Object.defineProperty(navigator, 'webdriver', { get: () => false });
    window.chrome = window.chrome || { runtime: {} };
    Object.defineProperty(navigator, 'languages', { get: () => ['ru-RU', 'ru', 'en-US', 'en'] });
    """

    def _url_candidates(src_url: str):
        out = [src_url]
        if "ru.dotabuff.com" in src_url:
            out.append(src_url.replace("ru.dotabuff.com", "www.dotabuff.com"))
        elif "www.dotabuff.com" in src_url:
            out.append(src_url.replace("www.dotabuff.com", "ru.dotabuff.com"))
        uniq = []
        for u in out:
            if u not in uniq:
                uniq.append(u)
        return uniq

    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(
                headless=not headed,
                channel="chrome",
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-infobars",
                    "--window-size=1920,1080",
                    "--disable-extensions",
                    "--disable-gpu",
                    "--disable-setuid-sandbox",
                ],
            )
        except Exception:
            browser = p.chromium.launch(
                headless=not headed,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-infobars",
                    "--window-size=1920,1080",
                ],
            )

        context_kwargs = {
            "user_agent": USER_AGENT,
            "locale": "ru-RU",
            "viewport": {"width": 1920, "height": 1080},
            "timezone_id": "Europe/Moscow",
            "extra_http_headers": {"Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8"},
        }
        if _playwright_storage_state:
            context_kwargs["storage_state"] = _playwright_storage_state
        context = browser.new_context(**context_kwargs)
        context.add_init_script(stealth_script)
        page = context.new_page()

        try:
            last_err = None
            loaded_url = None
            for candidate in _url_candidates(url):
                for attempt in range(1, 4):
                    try:
                        page.goto(candidate, wait_until="domcontentloaded", timeout=_clamp_timeout(timeout))
                        page.wait_for_load_state("domcontentloaded")
                        page.wait_for_timeout(_clamp_timeout(1500))
                        title_l = (page.title() or "").lower()
                        if "just a moment" in title_l:
                            raise RuntimeError("Cloudflare challenge page")
                        loaded_url = candidate
                        break
                    except Exception as e:
                        last_err = e
                        page.wait_for_timeout(_clamp_timeout(700 * attempt))
                if loaded_url:
                    break

            if not loaded_url:
                raise last_err if last_err else RuntimeError(f"Не удалось открыть URL: {url}")

            # Дать странице прогрузить основные блоки
            page.wait_for_timeout(_clamp_timeout(3000))

            # Ждём наиболее стабильные маркеры контента матча
            for selector in [
                ".match-victory-subtitle",
                ".the-radiant-score",
                ".the-dire-score",
                ".match-victory-subtitle .the-radiant.score",
                ".match-victory-subtitle .the-dire.score",
                "a[href*='/heroes/']",
                "table",
                ".content",
                "main",
            ]:
                try:
                    page.wait_for_selector(selector, timeout=_clamp_timeout(15000))
                    break
                except Exception:
                    continue

            page.wait_for_timeout(_clamp_timeout(int(random.uniform(1200, 2400))))
            html = None
            for _ in range(5):
                try:
                    html = page.content()
                    if html:
                        break
                except Exception:
                    page.wait_for_timeout(_clamp_timeout(800))
            if "Just a moment..." in html and "security verification" in html.lower():
                raise RuntimeError("Cloudflare challenge page after load")

            if save_debug_path:
                Path(save_debug_path).write_text(html or "", encoding="utf-8")

            try:
                _playwright_storage_state = context.storage_state()
            except Exception:
                pass

            return html
        finally:
            browser.close()


# ---------- OpenDota ----------
_heroes_cache = None
_abilities_cache = None

def get_heroes_map():
    global _heroes_cache
    if _heroes_cache is not None:
        return _heroes_cache
    try:
        r = requests.get(f"{OPENDOTA_URL}/constants/heroes", timeout=20)
        r.raise_for_status()
        data = r.json()
        out = {}
        for hid, h in data.items():
            out[int(hid)] = h.get("localized_name") or h.get("name", "?")
        _heroes_cache = out
        return out
    except Exception:
        _heroes_cache = {}
        return _heroes_cache

def get_abilities_map():
    global _abilities_cache
    if _abilities_cache is not None:
        return _abilities_cache
    try:
        r = requests.get(f"{OPENDOTA_URL}/constants/abilities", timeout=25)
        r.raise_for_status()
        data = r.json()
        out = {}
        for ability_name, meta in data.items():
            aid = meta.get("id")
            if aid is None:
                continue
            out[int(aid)] = {
                "name": ability_name,
                "dname": meta.get("dname") or "",
                "desc": meta.get("desc") or "",
                "hint": " / ".join(meta.get("hint") or []) if isinstance(meta.get("hint"), list) else "",
            }
        _abilities_cache = out
        return out
    except Exception:
        _abilities_cache = {}
        return _abilities_cache

LANE_MAP = {0: "?", 1: "Легкая", 2: "Средняя", 3: "Сложная"}
# lane_role OpenDota: 1 carry, 2 mid, 3 off, 4 soft, 5 hard
ROLE_MAP = {0: "?", 1: "Ключевая роль", 2: "Ключевая роль", 3: "Ключевая роль", 4: "Поддержка", 5: "Поддержка"}

def fetch_opendota_match(match_id: str):
    try:
        r = requests.get(f"{OPENDOTA_URL}/matches/{match_id}", timeout=20)
        if r.status_code != 200:
            return None
        return r.json()
    except Exception:
        return None

def _norm_hero_name(s: str) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    s = re.sub(r"[^a-zа-я0-9]+", "", s, flags=re.I)
    return s

def detect_position(lane: str, role: str) -> str:
    lane_l = (lane or "").lower()
    role_l = (role or "").lower()

    if ("ключ" in role_l or "core" in role_l) and ("легк" in lane_l or "safe" in lane_l):
        return "1"
    if ("ключ" in role_l or "core" in role_l) and ("сред" in lane_l or "mid" in lane_l):
        return "2"
    if ("ключ" in role_l or "core" in role_l) and ("слож" in lane_l or "off" in lane_l):
        return "3"
    if ("поддерж" in role_l or "support" in role_l) and ("слож" in lane_l or "off" in lane_l):
        return "4"
    if ("поддерж" in role_l or "support" in role_l) and ("легк" in lane_l or "safe" in lane_l):
        return "5"
    return ""

def detect_position_from_lane_role_value(lane_role_value) -> str:
    """
    OpenDota lane_role: 1 carry, 2 mid, 3 offlane, 4 soft support, 5 hard support.
    """
    if lane_role_value is None:
        return ""
    try:
        role_num = int(lane_role_value)
        if 1 <= role_num <= 5:
            return str(role_num)
    except Exception:
        pass

    role_str = str(lane_role_value).strip()
    m = re.match(r"^\s*([1-5])\b", role_str)
    return m.group(1) if m else ""

def fill_missing_positions(players: list):
    """
    Final fallback to avoid empty position values.
    Strategy: keep existing positions, fill gaps inside each team by GPM rank.
    """
    if not players:
        return players

    teams = {"Radiant": [], "Dire": []}
    for p in players:
        team = p.get("team") or ""
        if team in teams:
            teams[team].append(p)

    for team_rows in teams.values():
        if not team_rows:
            continue

        used = set()
        for r in team_rows:
            pos = (r.get("position") or "").strip()
            if pos in ("1", "2", "3", "4", "5"):
                used.add(pos)

        missing_positions = [x for x in ("1", "2", "3", "4", "5") if x not in used]
        rows_missing = [r for r in team_rows if (r.get("position") or "").strip() not in ("1", "2", "3", "4", "5")]
        if not rows_missing:
            continue

        def gpm_key(row):
            val = str(row.get("GPM") or "").replace(",", "").strip()
            m = re.search(r"\d+", val)
            return int(m.group(0)) if m else -1

        rows_missing.sort(key=gpm_key, reverse=True)

        for i, row in enumerate(rows_missing):
            if i < len(missing_positions):
                row["position"] = missing_positions[i]
            else:
                # safety fallback, should not happen for normal 5-player team rows
                row["position"] = row.get("position") or "5"

    return players

def enrich_players_with_opendota(match_id: str, players: list, heroes_map: dict):
    match_data = fetch_opendota_match(match_id)
    if not match_data or "players" not in match_data:
        return players

    abilities_map = get_abilities_map()
    od_players = match_data.get("players") or []

    by_team_hero = {}
    by_hero_only = {}
    for p in od_players:
        slot = p.get("player_slot", 0)
        team = "Radiant" if slot < 128 else "Dire"
        hero_id = p.get("hero_id")
        hero_name = heroes_map.get(hero_id, str(hero_id)) if hero_id else "?"
        hero_norm = _norm_hero_name(hero_name)
        by_team_hero[(team, hero_norm)] = p
        by_hero_only[hero_norm] = p

    for row in players:
        team = row.get("team") or ""
        hero = row.get("hero") or ""
        hero_norm = _norm_hero_name(hero)
        p = by_team_hero.get((team, hero_norm)) or by_hero_only.get(hero_norm)
        if not p:
            continue

        if not (row.get("lane") or "").strip() or row.get("lane") in ("?", "Safe", "Mid", "Off") or any(x in (row.get("lane") or "").lower() for x in ("роум", "roam", "jungle", "лес")):
            row["lane"] = LANE_MAP.get(p.get("lane", 0), "?")
        if not (row.get("lane_role") or "").strip() or row.get("lane_role") in ("?", "1 (Carry)", "2 (Mid)", "3 (Offlane)", "4 (Soft Support)", "5 (Hard Support)"):
            row["lane_role"] = ROLE_MAP.get(p.get("lane_role", 0), "?")

        abil_arr = p.get("ability_upgrades_arr") or []
        talent_id = abil_arr[9] if len(abil_arr) > 9 else ""
        row["first_talent_id"] = str(talent_id) if talent_id else ""

        if talent_id and int(talent_id) in abilities_map:
            meta = abilities_map[int(talent_id)]
            text = meta.get("dname") or meta.get("name") or ""
            row["first_talent"] = text
            row["first_talent_text"] = text
        else:
            row["first_talent"] = row.get("first_talent") or ""
            row["first_talent_text"] = row.get("first_talent_text") or ""

        pos_from_lane_role = detect_position_from_lane_role_value(p.get("lane_role"))
        if pos_from_lane_role:
            row["position"] = pos_from_lane_role
        else:
            row["position"] = row.get("position") or detect_position(row.get("lane", ""), row.get("lane_role", ""))

    # score/winner/duration можно тоже брать отсюда, но это делаем в parse_overview
    return players


# ---------- Dotabuff HTML parsing ----------
HEADER_TO_KEY = {
    "K": "K", "D": "D", "A": "A", "L": "L", "R": "R", "Л": "L", "Р": "R",
    "GPM": "GPM", "XPM": "XPM", "LH": "LH", "DN": "DN",
    "У": "K", "С": "D", "П": "A",
    "ОЦ": "GPM", "Опыт": "XPM", "Добито": "LH", "Зап.": "DN", "Зап": "DN",
    "Lane": "Lane", "Role": "Role", "Линия": "Lane", "Роль": "Role",
    "Hero": "Hero", "Герой": "Hero", "HERO": "Hero",
}

def extract_first_talent_text_from_dotabuff_skillbuild(soup, hero_name):
    """
    Парс выбранного таланта на 10 уровне из Dotabuff skillbuild-table (overview HTML).
    В talents-ячейке есть скрытый <div style="display:none"><table>...> где выбранный талант помечен
    классом 'talent-cell active' в строке уровня 10.
    Возвращает текст таланта или "".
    """
    if not soup or not hero_name:
        return ""

    hero_norm = hero_name.strip().lower().replace(" ", "").replace("'", "")
    for tr in soup.select("tr"):
        talents_td = tr.select_one("td.ability.talents")
        if not talents_td:
            continue

        a = tr.select_one("a[href*='/heroes/']")
        if not a:
            continue

        href = (a.get("href") or "").lower()
        txt = (a.get_text(strip=True) or "").lower()
        slug = href.split("/heroes/")[-1].split("/")[0].replace("-", "")
        txtn = txt.replace(" ", "").replace("'", "")
        if hero_norm not in slug and hero_norm not in txtn:
            continue

        hidden = talents_td.select_one("div[style*='display'] table")
        if not hidden:
            continue

        for row in hidden.select("tbody tr"):
            lvl = row.select_one(".talent-level")
            if not lvl:
                continue
            if lvl.get_text(" ", strip=True) != "10":
                continue

            # chosen is marked as active
            active = row.select_one("td.talent-cell.active") or row.select_one("td.talent-cell.is-active") or row.select_one("td.talent-cell.selected")
            if active:
                return active.get_text(" ", strip=True)

            actives = row.select("td.talent-cell.active")
            if actives:
                return actives[0].get_text(" ", strip=True)

            return ""

    return ""


def _hero_name_from_href(href):
    if not href:
        return ""
    m = re.search(r"/heroes/([^/?#]+)", href)
    if not m:
        return ""
    return m.group(1).replace("-", " ").title()

def _header_to_key(header_text):
    if not header_text:
        return None
    h = header_text.strip()
    if h in HEADER_TO_KEY:
        return HEADER_TO_KEY[h]
    h_lower = h.lower()
    if "lane" in h_lower or "линия" in h_lower:
        return "Lane"
    if "role" in h_lower or "роль" in h_lower:
        return "Role"
    if "hero" in h_lower or "герой" in h_lower:
        return "Hero"
    if "kill" in h_lower or "убийств" in h_lower:
        return "K"
    if "death" in h_lower or "смерт" in h_lower:
        return "D"
    if "assist" in h_lower or "помощ" in h_lower:
        return "A"
    if "gpm" in h_lower or "золото" in h_lower or h in ("ОЦ", "NET"):
        return "GPM"
    if "xpm" in h_lower or "опыт" in h_lower:
        return "XPM"
    if "last" in h_lower or "добито" in h_lower or "lh" in h_lower:
        return "LH"
    if "den" in h_lower or "зап" in h_lower or "dn" in h_lower:
        return "DN"
    return None

def parse_overview(html: str, match_id: str):
    if not html:
        return {"match_id": match_id, "error": "Нет HTML", "players": [], "players_count": 0, "duration": "", "score": "", "winner": ""}

    soup = BeautifulSoup(html, "html.parser")
    result = {"match_id": match_id, "duration": "", "score": "", "winner": ""}

    # ---- duration (точно) ----
    dur_el = soup.select_one(".duration")
    if dur_el:
        result["duration"] = dur_el.get_text(strip=True)

    # ---- score (точно из DOM, с fallback по regex на случай иной вёрстки) ----
        # В разных версиях разметки встречается:
    # 1) .the-radiant-score / .the-dire-score
    # 2) <span class="the-radiant score"> / <span class="the-dire score">
    r_el = (soup.select_one(".the-radiant-score")
            or soup.select_one(".match-victory-subtitle .the-radiant.score")
            or soup.select_one(".match-victory-subtitle span.the-radiant.score"))
    d_el = (soup.select_one(".the-dire-score")
            or soup.select_one(".match-victory-subtitle .the-dire.score")
            or soup.select_one(".match-victory-subtitle span.the-dire.score"))
    r_score = r_el.get_text(strip=True) if r_el else ""
    d_score = d_el.get_text(strip=True) if d_el else ""

    if not (r_score and d_score):
        # fallback: ищем прямо в html (иногда элементы есть, но soup не видит из-за неполной отрисовки)
        m = re.search(r'the-radiant-score[^>]*>\s*(\d+)\s*<.*?the-dire-score[^>]*>\s*(\d+)\s*<', html, flags=re.S|re.I)
        if m:
            r_score, d_score = m.group(1), m.group(2)

    if r_score and d_score:
        result["score"] = f"{r_score} - {d_score}"
        try:
            r_i, d_i = int(r_score), int(d_score)
            result["winner"] = "Radiant" if r_i > d_i else "Dire" if d_i > r_i else ""
        except Exception:
            pass

    # ---- winner fallback by text (если вдруг score не взялся) ----
    if not result["winner"]:
        full_text = soup.get_text(" ", strip=True)
        if re.search(r"Победа\s+сил\s+Света|Победа\s+Света|Radiant\s+Victory", full_text, re.I):
            result["winner"] = "Radiant"
        elif re.search(r"Победа\s+сил\s+Тьмы|Победа\s+Тьмы|Dire\s+Victory", full_text, re.I):
            result["winner"] = "Dire"

    # ---- players table detection ----
    def _tooltip(cell):
        # Dotabuff часто кладёт подсказку в oldtitle (а title пустой)
        return (
            cell.get("title")
            or cell.get("data-original-title")
            or cell.get("data-tooltip")
            or cell.get("oldtitle")
            or cell.get("aria-label")
            or ""
        ).strip()

    def _team_from_context(el):
        p = el
        for _ in range(20):
            if not p:
                break
            t = (p.get_text(" ", strip=True) or "").lower()
            c = " ".join(p.get("class", [])).lower() if hasattr(p, "get") else ""
            if "radiant" in c or "силы света" in t:
                return "Radiant"
            if "dire" in c or "силы тьмы" in t:
                return "Dire"
            p = p.parent
        return None

    def _is_players_table(table):
        thead = table.find("thead")
        head_row = thead.find("tr") if thead else table.find("tr")
        if not head_row:
            return False
        headers = [th.get_text(strip=True) for th in head_row.find_all(["th", "td"])]
        headers_norm = " ".join([h.upper() for h in headers])
        if not any(x in headers_norm for x in ["K", "D", "A", "У", "С", "П", "L", "R", "Л", "Р"]):
            return False
        tbody = table.find("tbody") or table
        hero_links = tbody.find_all("a", href=re.compile(r"/heroes/"))
        return len(hero_links) >= 5

    def _parse_table(table):
        out = []
        thead = table.find("thead")
        tbody = table.find("tbody") or table
        head_row = thead.find("tr") if thead else table.find("tr")
        headers_raw = []
        for th in head_row.find_all(["th", "td"]):
            txt = (th.get_text(strip=True) or "").strip()
            tt = (th.get("title") or th.get("data-original-title") or th.get("data-tooltip") or th.get("oldtitle") or th.get("aria-label") or "").lower()
            if "lane" in tt or "линия" in tt:
                headers_raw.append("Lane")
            elif "role" in tt or "роль" in tt:
                headers_raw.append("Role")
            else:
                headers_raw.append(txt)

        header_keys = [_header_to_key(h) for h in headers_raw]
        table_team = _team_from_context(table)

        for tr in tbody.find_all("tr"):
            cells = tr.find_all(["td", "th"])
            if len(cells) < 2:
                continue
            row = {"match_id": match_id}
            if table_team:
                row["team"] = table_team

            for i, cell in enumerate(cells):
                if i >= len(header_keys) or not header_keys[i]:
                    continue
                key = header_keys[i]

                a = cell.find("a", href=re.compile(r"/heroes/"))
                if a:
                    hero_text = a.get_text(strip=True)
                    row["hero"] = hero_text if hero_text and not hero_text.isdigit() else _hero_name_from_href(a.get("href", "")) or hero_text
                    facet_img = cell.select_one("img[src*='/assets/facet_icons/']")
                    if facet_img:
                        facet_text = (
                            facet_img.get("alt")
                            or facet_img.get("title")
                            or facet_img.get("data-original-title")
                            or facet_img.get("aria-label")
                            or ""
                        ).strip()
                        if facet_text:
                            row["start_talent"] = facet_text

                val = cell.get_text(strip=True)
                if (not val) and cell.get("data-value") is not None:
                    val = str(cell.get("data-value")).strip()

                # L/R часто только в tooltip
                if key in ("Lane", "Role", "L", "R"):
                    # В ячейке обычно лежит <i ... oldtitle="Средняя линия"> (подсказка не на td)
                    tt = _tooltip(cell).lower()
                    if tt:
                        val = tt
                    else:
                        inner = (cell.find(attrs={"title": True})
                                 or cell.find(attrs={"data-original-title": True})
                                 or cell.find(attrs={"data-tooltip": True})
                                 or cell.find(attrs={"oldtitle": True})
                                 or cell.find(attrs={"aria-label": True}))
                        if inner:
                            val = _tooltip(inner).lower() or val

                row[key] = val

            if row.get("hero"):
                out.append(row)
        return out

    players = []
    for table in soup.find_all("table"):
        if _is_players_table(table):
            players.extend(_parse_table(table))

    # normalise
    for p in players:
        p["lane"] = (p.get("L") or p.get("Lane") or "").strip()
        p["lane_role"] = (p.get("R") or p.get("Role") or "").strip()

    # dedupe (team+hero)
    unique = []
    seen = set()
    for p in players:
        hero = (p.get("hero") or "").strip()
        if not hero:
            continue
        team = p.get("team") or ""
        k = (team, _norm_hero_name(hero))
        if k in seen:
            continue
        seen.add(k)
        unique.append(p)

    # team fallback by order
    for idx, p in enumerate(unique):
        if not p.get("team"):
            p["team"] = "Radiant" if idx < 5 else "Dire"

    unique = unique[:10]

    PLAYER_KEYS = ("match_id","team","hero","lane","lane_role","position","start_talent","first_talent_id","first_talent","first_talent_text","K","D","A","GPM","XPM","LH","DN")
    clean = []
    for p in unique:
        row = {k: "" for k in PLAYER_KEYS}
        row["match_id"] = match_id
        row["team"] = p.get("team","")
        row["hero"] = (p.get("hero") or "").strip()
        row["lane"] = (p.get("lane") or "").strip()
        row["lane_role"] = (p.get("lane_role") or "").strip()
        row["start_talent"] = (p.get("start_talent") or "").strip()
        for k in ("K","D","A","GPM","XPM","LH","DN"):
            v = p.get(k)
            row[k] = (v or "").strip() if isinstance(v,str) else (str(v) if v is not None else "")
        row["position"] = detect_position_from_lane_role_value(row["lane_role"]) or detect_position(row["lane"], row["lane_role"])
        clean.append(row)

    # enrich from OpenDota (lane/role/position + talents)
    try:
        clean = enrich_players_with_opendota(match_id, clean, heroes_map=get_heroes_map())
    except Exception:
        pass

    # Dotabuff fallback: если текст таланта не пришёл из OpenDota — берём из skillbuild таблицы (lvl10 chosen)
    try:
        for row in clean:
            if (row.get("start_talent") or "").strip():
                row["first_talent"] = row["start_talent"]
                row["first_talent_text"] = row["start_talent"]

            if not (row.get("first_talent_text") or "").strip():
                txt = extract_first_talent_text_from_dotabuff_skillbuild(soup, row.get("hero",""))
                if txt:
                    row["first_talent_text"] = txt
                    if not (row.get("first_talent") or "").strip():
                        row["first_talent"] = txt
    except Exception:
        pass

    try:
        clean = fill_missing_positions(clean)
    except Exception:
        pass

    result["players"] = clean
    result["players_count"] = len(clean)
    return result


def parse_lanes_tab(html, match_id):
    if not html:
        return []
    soup = BeautifulSoup(html, "html.parser")

    def _text_or_dataval(cell):
        v = cell.get_text(" ", strip=True)
        if (not v) and cell.get("data-value") is not None:
            v = str(cell.get("data-value")).strip()
        return v

    def _split_lane_cell(cell):
        parts = [s.strip() for s in cell.stripped_strings if s.strip()]
        joined = " ".join(parts)
        outcome = ""
        jl = joined.lower()
        if "проигрыш" in jl:
            outcome = "Проигрыш"
        elif "выигрыш" in jl or "победа" in jl:
            outcome = "Выигрыш"
        elif "ничья" in jl:
            outcome = "Ничья"

        team = ""
        if "силы света" in jl:
            team = "Силы Света"
        elif "силы тьмы" in jl:
            team = "Силы Тьмы"

        detail = ""
        if parts:
            cand = [p for p in parts if "линия" not in p.lower() and "силы" not in p.lower() and p.lower() not in ("проигрыш","выигрыш","ничья")]
            detail = cand[-1] if cand else parts[-1]

        lane_simple = ""
        side = ""
        dl = detail.lower()

        # lane type: handle full words + common abbreviations from Dotabuff.
        if ("легк" in dl) or ("safe" in dl) or re.search(r"\bлег\b", dl):
            lane_simple = "Легкая"
        elif ("сред" in dl) or ("мид" in dl) or ("mid" in dl) or re.search(r"\bсрд\b", dl):
            lane_simple = "Средняя"
        elif ("слож" in dl) or ("off" in dl) or re.search(r"\bслж\b", dl):
            lane_simple = "Сложная"
        elif ("роум" in dl) or ("roam" in dl):
            lane_simple = "Роуминг"

        # lane side: top/bot; mid is center.
        if ("верх" in dl) or ("top" in dl):
            side = "Верхняя"
        elif ("ниж" in dl) or ("bot" in dl):
            side = "Нижняя"
        elif lane_simple == "Средняя":
            side = "Центр"
        elif lane_simple == "Роуминг":
            side = "Роум"

        # extra fallback for patterns like "Слж/Верхняя" or "Safe/Top"
        if not side and "/" in dl:
            right = dl.split("/")[-1].strip()
            if ("верх" in right) or ("top" in right):
                side = "Верхняя"
            elif ("ниж" in right) or ("bot" in right):
                side = "Нижняя"

        return outcome, team, detail, lane_simple, side

    def _first_int_like(text):
        if not text:
            return ""
        m = re.search(r"-|[\d][\d,\s]*", text)
        if not m:
            return ""
        v = m.group(0).strip()
        if v == "-":
            return "-"
        return re.sub(r"[^\d]", "", v)

    def _cell_primary_value(cell):
        if cell is None:
            return ""
        grp = cell.select_one(".group")
        if grp:
            txt = grp.get_text(" ", strip=True)
            return _first_int_like(txt)
        return _first_int_like(cell.get_text(" ", strip=True))

    # Modern lanes table parser (current Dotabuff layout).
    rows = []
    for tr in soup.find_all("tr"):
        cls = " ".join(tr.get("class", []))
        if "player-" not in cls or "faction-" not in cls:
            continue

        row = {"match_id": match_id}

        hero_a = tr.select_one("td.r-tab-icon a[href*='/heroes/']")
        hero_img = tr.select_one("td.r-tab-icon img")
        hero_name = ""
        if hero_a:
            hero_name = _hero_name_from_href(hero_a.get("href", "")) or ""
        if not hero_name and hero_img:
            src = (hero_img.get("src") or "").strip()
            m_slug = re.search(r"/(?:miniheroes|heroes)/([^/.]+)", src)
            if m_slug:
                hero_name = m_slug.group(1).replace("-", " ").title()
        if not hero_name and hero_img:
            hero_name = (
                hero_img.get("oldtitle")
                or hero_img.get("title")
                or hero_img.get("alt")
                or ""
            ).strip()
        row["hero"] = hero_name

        lane_cell = None
        for c in tr.find_all("td", class_=re.compile(r"\bcell-divider\b")):
            t = c.get_text(" ", strip=True).lower()
            if "линия" in t or "line" in t:
                lane_cell = c
                break
        if lane_cell is None:
            lane_cell = tr.find("td", class_=re.compile(r"\bcell-divider\b"))
        if lane_cell is not None:
            outcome, team, detail, lane_simple, side = _split_lane_cell(lane_cell)
            row["lane_outcome"] = outcome
            row["lane_team"] = team
            row["lane_detail"] = detail
            row["lane_simple"] = lane_simple
            row["lane_side"] = side
            row["lane"] = lane_simple or detail

        row["gpm_12"] = _cell_primary_value(tr.find("td", class_=re.compile(r"\bcolor-stat-gold\b")))
        row["xpm_12"] = _cell_primary_value(tr.find("td", class_=re.compile(r"\bcolor-stat-experience\b")))
        row["k_12"] = _cell_primary_value(tr.find("td", class_=re.compile(r"\bcolor-stat-kill\b")))
        row["d_12"] = _cell_primary_value(tr.find("td", class_=re.compile(r"\bcolor-stat-death\b")))
        row["a_12"] = _cell_primary_value(tr.find("td", class_=re.compile(r"\bcolor-stat-assist\b")))

        centered = tr.find_all("td", class_=re.compile(r"\bcell-centered\b"))
        if len(centered) >= 3:
            row["lh_4"] = _first_int_like(centered[-3].get_text(" ", strip=True))
            row["lh_8"] = _first_int_like(centered[-2].get_text(" ", strip=True))
            row["lh_12"] = _first_int_like(centered[-1].get_text(" ", strip=True))

        if row.get("hero"):
            rows.append(row)

    if rows:
        return rows

    rows = []
    for table in soup.find_all("table"):
        thead = table.find("thead")
        tbody = table.find("tbody")
        if not thead or not tbody:
            continue

        headers = []
        for th in thead.find_all(["th","td"]):
            htxt = th.get_text(" ", strip=True)
            htitle = (th.get("title") or th.get("data-original-title") or th.get("data-tooltip") or "").strip()
            headers.append(htxt or htitle)

        head_join = " ".join(headers).lower()
        hero_links = len(tbody.select("a[href*='/heroes/']"))
        RUS_HDR_LANE = "\u043b\u0438\u043d\u0438\u044f"
        RUS_HDR_HERO = "\u0433\u0435\u0440\u043e\u0439"
        RUS_HDR_KILLS = "\u0443\u0431"
        RUS_HDR_DEATHS = "\u0441\u043c\u0435\u0440\u0442"
        RUS_HDR_ASSISTS = "\u043f\u043e\u043c\u043e\u0449"
        RUS_HDR_GPM = "\u0437\u043e\u043b\u043e\u0442"
        RUS_HDR_XPM = "\u043e\u043f\u044b\u0442"

        has_lane_like_headers = any(
            k in head_join for k in (
                "lane", "hero", "gpm", "xpm", "net", "@", "kill", "death", "assist",
                RUS_HDR_LANE, RUS_HDR_HERO, RUS_HDR_KILLS, RUS_HDR_DEATHS, RUS_HDR_ASSISTS,
                RUS_HDR_GPM, RUS_HDR_XPM
            )
        )
        if hero_links < 5 and not has_lane_like_headers:
            continue

        keys = []
        for h in headers:
            h = (h or "").strip()
            hl = h.lower()
            if h in ("Hero",) or "hero" in hl or RUS_HDR_HERO in hl:
                keys.append("hero"); continue
            if h in ("Lane",) or "lane" in hl or RUS_HDR_LANE in hl:
                keys.append("lane_cell"); continue
            if h in ("GPM", "Net") or "gpm" in hl or "net" in hl or RUS_HDR_GPM in hl:
                keys.append("gpm_12"); continue
            if h.lower() in ("xpm",) or "xpm" in hl or RUS_HDR_XPM in hl:
                keys.append("xpm_12"); continue
            if h in ("K",) or "kill" in hl or RUS_HDR_KILLS in hl:
                keys.append("k_12"); continue
            if h in ("D",) or "death" in hl or RUS_HDR_DEATHS in hl:
                keys.append("d_12"); continue
            if h in ("A",) or "assist" in hl or RUS_HDR_ASSISTS in hl:
                keys.append("a_12"); continue
            m = re.search(r"@\s*(4|8|12)", h) or re.search(r"(?:^|\D)(4|8|12)(?:\D|$)", h)
            if m:
                keys.append(f"lh_{m.group(1)}"); continue
            keys.append(h or None)

        for tr in tbody.find_all("tr"):
            cells = tr.find_all(["td","th"])
            if not cells:
                continue
            row = {"match_id": match_id}
            for i, cell in enumerate(cells):
                if i >= len(keys): break
                key = keys[i]
                if not key: continue

                a = cell.find("a", href=re.compile(r"/heroes/"))
                if a and key in ("hero","lane_cell"):
                    hn = a.get_text(strip=True)
                    row["hero"] = hn if hn and not hn.isdigit() else _hero_name_from_href(a.get("href","")) or hn

                if key == "lane_cell":
                    outcome, team, detail, lane_simple, side = _split_lane_cell(cell)
                    row["lane_outcome"] = outcome
                    row["lane_team"] = team
                    row["lane_detail"] = detail
                    row["lane_simple"] = lane_simple
                    row["lane_side"] = side
                    row["lane"] = lane_simple or detail
                else:
                    row[key] = _text_or_dataval(cell)

            if row.get("hero"):
                rows.append(row)

    return rows


def load_match_ids(csv_path):
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {csv_path}")
    df = pd.read_csv(path)
    if "match_id" in df.columns:
        return df["match_id"].astype(str).str.strip().tolist()
    return df.iloc[:, 0].astype(str).str.strip().tolist()

WINNER_COL = "\u043f\u043e\u0431\u0435\u0434\u0438\u0442\u0435\u043b\u044c"
RUS_SHEET_SUMMARY = "\u041c\u0430\u0442\u0447\u0438_\u0441\u0432\u043e\u0434\u043a\u0430"
RUS_SHEET_PLAYERS = "\u0418\u0433\u0440\u043e\u043a\u0438_\u043e\u0431\u0437\u043e\u0440"
RUS_SHEET_LANES = "\u041b\u0438\u043d\u0438\u0438"
RUS_SHEET_LANING = "\u041b\u0430\u0439\u043d\u0438\u043d\u0433_\u0441\u0442\u0430\u0442\u0438\u0441\u0442\u0438\u043a\u0430"
RUS_SHEET_LANING_ALL = "\u041b\u0430\u0439\u043d\u0438\u043d\u0433_\u0432\u0441\u044f_\u0441\u0442\u0430\u0442\u0438\u0441\u0442\u0438\u043a\u0430"

DEFAULT_SHEETS = {
    "summary": "Matches_Summary",
    "players": "Players_Overview",
    "lanes": "Lanes_Raw",
    "laning": "Laning_Stats",
    "laning_all": "Laning_All",
}

SHEET_CANDIDATES = {
    "summary": [RUS_SHEET_SUMMARY, DEFAULT_SHEETS["summary"]],
    "players": [RUS_SHEET_PLAYERS, DEFAULT_SHEETS["players"]],
    "lanes": [RUS_SHEET_LANES, DEFAULT_SHEETS["lanes"]],
    "laning": [RUS_SHEET_LANING, DEFAULT_SHEETS["laning"]],
    "laning_all": [RUS_SHEET_LANING_ALL, DEFAULT_SHEETS["laning_all"]],
}

SUMMARY_COLS = ["match_id", "duration", "score", WINNER_COL, "winner", "players_parsed", "error"]
PLAYER_COLS = [
    "match_id", "team", "hero", "lane", "lane_role", "position",
    "start_talent", "first_talent_id", "first_talent", "first_talent_text",
    "K", "D", "A", "GPM", "XPM", "LH", "DN"
]
LANES_RAW_COLS = [
    "match_id", "hero", "lane_outcome", "lane_team", "lane_detail", "lane_simple",
    "lane_side", "lane", "gpm_12", "xpm_12", "k_12", "d_12", "a_12", "lh_4", "lh_8", "lh_12"
]
LANING_COLS = [
    "match_id", "hero", "lane_outcome", "lane_team", "lane_detail",
    "lane_simple", "lane_side", "gpm_12", "xpm_12",
    "kills_12", "deaths_12", "assists_12", "lh_4", "lh_8", "lh_12"
]

def init_logging(log_path):
    logger = logging.getLogger("dotabuff")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)
    log_path = Path(log_path)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(fmt)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(fmt)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger

def _safe_sheet_title(title):
    if title is None:
        return ""
    invalid = set("[]:*?/\\")
    cleaned = "".join(ch for ch in str(title) if ch not in invalid)
    cleaned = cleaned[:31]
    return cleaned or "Sheet"


def _resolve_sheet_name(wb, logical):
    for name in SHEET_CANDIDATES.get(logical, []):
        if name in wb.sheetnames:
            return name
    return DEFAULT_SHEETS.get(logical, logical)


def _get_or_create_sheet(wb, sheet_name, headers):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        safe_name = _safe_sheet_title(sheet_name)
        if safe_name in wb.sheetnames:
            ws = wb[safe_name]
        else:
            ws = wb.create_sheet(safe_name)

    existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not any(existing):
        ws.append(headers)
        return ws, headers

    headers_clean = [str(h) if h is not None else "" for h in existing]
    return ws, headers_clean


def _get_sheet_by_logical(wb, logical, headers):
    name = _resolve_sheet_name(wb, logical)
    ws, hdrs = _get_or_create_sheet(wb, name, headers)
    return ws, hdrs, name


def _append_rows(ws, rows, headers):
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

def _read_existing_match_ids(ws, headers):
    if not ws or not headers:
        return set()
    try:
        idx = headers.index("match_id") + 1
    except ValueError:
        return set()
    ids = set()
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
        val = row[0]
        if val is None:
            continue
        ids.add(str(val).strip())
    return ids

def _find_rows_by_match_id(ws, headers, match_id):
    if not ws or not headers:
        return []
    try:
        idx = headers.index("match_id") + 1
    except ValueError:
        return []
    out = []
    for row in ws.iter_rows(min_row=2):
        cell = row[idx - 1]
        val = cell.value
        if val is None:
            continue
        if str(val).strip() == str(match_id).strip():
            out.append((cell.row, [c.value for c in row]))
    return out

def _row_to_dict(headers, values):
    out = {}
    for i, h in enumerate(headers):
        if i >= len(values):
            out[h] = ""
        else:
            out[h] = values[i]
    return out

def _delete_rows(ws, row_indexes):
    if not row_indexes:
        return 0
    count = 0
    for r in sorted(set(row_indexes), reverse=True):
        ws.delete_rows(r)
        count += 1
    return count

def _delete_match_rows(ws, headers, match_id):
    rows = _find_rows_by_match_id(ws, headers, match_id)
    return _delete_rows(ws, [r for r, _ in rows])

def _dedup_sheet(ws, headers, key_cols):
    if not ws or not headers:
        return 0
    idxs = []
    for c in key_cols:
        if c in headers:
            idxs.append(headers.index(c) + 1)
    if not idxs:
        return 0
    seen = set()
    to_delete = []
    for row in ws.iter_rows(min_row=2):
        key = []
        for idx in idxs:
            val = row[idx - 1].value
            key.append(str(val).strip() if val is not None else "")
        if not any(key):
            continue
        key = tuple(key)
        if key in seen:
            to_delete.append(row[0].row)
        else:
            seen.add(key)
    return _delete_rows(ws, to_delete)

def _is_missing_value(val):
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s == "?"


def _is_summary_row_complete(row):
    if not row:
        return False
    duration = row.get("duration")
    score = row.get("score")
    winner = row.get("winner")
    winner_label = row.get(WINNER_COL)
    error = row.get("error")
    players_parsed = row.get("players_parsed")
    try:
        players_parsed = int(str(players_parsed).strip()) if players_parsed is not None else 0
    except Exception:
        players_parsed = 0

    if error and str(error).strip():
        return False
    if _is_missing_value(duration) or _is_missing_value(score):
        return False
    if _is_missing_value(winner) and _is_missing_value(winner_label):
        return False
    if players_parsed and players_parsed < 10:
        return False
    return True


def _count_rows_by_match_id(ws, headers, require_hero=True):
    if not ws or not headers:
        return {}
    try:
        idx_mid = headers.index("match_id")
    except ValueError:
        return {}
    idx_hero = None
    if require_hero:
        try:
            idx_hero = headers.index("hero")
        except ValueError:
            idx_hero = None
    counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_mid >= len(row):
            continue
        mid = row[idx_mid]
        if mid is None:
            continue
        if require_hero and idx_hero is not None:
            if idx_hero >= len(row):
                continue
            hero = row[idx_hero]
            if hero is None or str(hero).strip() == "":
                continue
        key = str(mid).strip()
        counts[key] = counts.get(key, 0) + 1
    return counts

def _get_sheet_if_exists(wb, logical):
    for name in SHEET_CANDIDATES.get(logical, []):
        if name in wb.sheetnames:
            return wb[name], name
    return None, None


def _scan_sheet_counts(ws, headers, required_cols):
    if not ws or not headers:
        return {}, {}
    if "match_id" not in headers:
        return {}, {}
    idx_mid = headers.index("match_id")
    idxs_req = [headers.index(c) for c in required_cols if c in headers]
    counts = {}
    missing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_mid >= len(row):
            continue
        mid = row[idx_mid]
        if mid is None:
            continue
        key = str(mid).strip()
        counts[key] = counts.get(key, 0) + 1
        if missing.get(key, False):
            continue
        for idx in idxs_req:
            if idx >= len(row) or _is_missing_value(row[idx]):
                missing[key] = True
                break
        else:
            missing.setdefault(key, False)
    return counts, missing


def _scan_summary_sheet(ws, headers):
    if not ws or not headers:
        return {}, {}
    if "match_id" not in headers:
        return {}, {}
    idx_mid = headers.index("match_id")
    idx_duration = headers.index("duration") if "duration" in headers else None
    idx_score = headers.index("score") if "score" in headers else None
    idx_winner = headers.index("winner") if "winner" in headers else None
    idx_winner_label = headers.index(WINNER_COL) if WINNER_COL in headers else None
    idx_error = headers.index("error") if "error" in headers else None
    idx_players = headers.index("players_parsed") if "players_parsed" in headers else None

    counts = {}
    missing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_mid >= len(row):
            continue
        mid = row[idx_mid]
        if mid is None:
            continue
        key = str(mid).strip()
        counts[key] = counts.get(key, 0) + 1
        if missing.get(key, False):
            continue

        def _val(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        duration = _val(idx_duration)
        score = _val(idx_score)
        winner = _val(idx_winner)
        winner_label = _val(idx_winner_label)
        error = _val(idx_error)
        players_val = _val(idx_players)

        if error and str(error).strip():
            missing[key] = True
            continue
        if _is_missing_value(duration) or _is_missing_value(score):
            missing[key] = True
            continue
        if _is_missing_value(winner) and _is_missing_value(winner_label):
            missing[key] = True
            continue
        if players_val is not None:
            try:
                if int(str(players_val).strip()) < 10:
                    missing[key] = True
                    continue
            except Exception:
                missing[key] = True
                continue
        missing.setdefault(key, False)

    return counts, missing


def analyze_existing_excel(path, logger=None):
    if not Path(path).exists():
        return {}
    wb = load_workbook(path)

    summary_counts = {}
    summary_missing = {}
    ws, _ = _get_sheet_if_exists(wb, "summary")
    if ws is not None:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h) if h is not None else "" for h in headers]
        summary_counts, summary_missing = _scan_summary_sheet(ws, headers)

    players_counts = {}
    players_missing = {}
    ws, _ = _get_sheet_if_exists(wb, "players")
    if ws is not None:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h) if h is not None else "" for h in headers]
        players_counts, players_missing = _scan_sheet_counts(
            ws,
            headers,
            ["team", "hero", "lane", "lane_role", "position"]
        )

    lanes_counts = {}
    lanes_missing = {}
    ws, _ = _get_sheet_if_exists(wb, "lanes")
    if ws is not None:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h) if h is not None else "" for h in headers]
        lanes_counts, lanes_missing = _scan_sheet_counts(
            ws,
            headers,
            ["hero", "lane_outcome", "lane_detail", "lane_simple", "lane_side"]
        )

    status = {}
    all_ids = set(summary_counts.keys()) | set(players_counts.keys()) | set(lanes_counts.keys())
    for mid in all_ids:
        summary_ok = summary_counts.get(mid, 0) >= 1 and not summary_missing.get(mid, True)
        players_ok = players_counts.get(mid, 0) >= 10 and not players_missing.get(mid, True)
        lanes_ok = lanes_counts.get(mid, 0) >= 10 and not lanes_missing.get(mid, True)
        need_overview = not (summary_ok and players_ok)
        need_lanes = not lanes_ok
        status[mid] = {
            "summary_ok": summary_ok,
            "players_ok": players_ok,
            "lanes_ok": lanes_ok,
            "need_overview": need_overview,
            "need_lanes": need_lanes,
            "complete": (not need_overview and not need_lanes),
        }

    if logger:
        total = len(status)
        complete = sum(1 for v in status.values() if v.get("complete"))
        logger.info(f"Existing Excel status: {complete}/{total} complete")
    return status


def deduplicate_excel(path, logger=None):
    path = Path(path)
    if not path.exists():
        return 0
    wb = load_workbook(path)
    removed = 0

    for logical, key_cols in (
        ("summary", ["match_id"]),
        ("players", ["match_id", "team", "hero"]),
        ("lanes", ["match_id", "hero"]),
        ("laning", ["match_id", "hero"]),
        ("laning_all", ["match_id", "hero"]),
    ):
        ws, _ = _get_sheet_if_exists(wb, logical)
        if ws is None:
            continue
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h) if h is not None else "" for h in headers]
        removed += _dedup_sheet(ws, headers, key_cols)

    if removed > 0:
        wb.save(path)
    if logger:
        logger.info(f"Deduplicate Excel: removed {removed} rows")
    return removed


def update_match_in_excel(overview, lanes_rows, output_path, update_overview=True, update_lanes=True, logger=None):
    path = Path(output_path)
    if not path.exists():
        if not update_overview or overview is None:
            raise ValueError("Excel does not exist; overview required to create file")
        save_to_excel([overview], [lanes_rows or []], path)
        return True

    wb = load_workbook(path)

    mid = str(overview.get("match_id") or "").strip() if overview else ""
    if not mid and lanes_rows:
        try:
            mid = str((lanes_rows[0] or {}).get("match_id") or "").strip()
        except Exception:
            mid = ""

    if update_overview and overview is not None:
        winner_val = overview.get("winner") or ""
        winner_label = "Radiant (\u0421\u0432\u0435\u0442\u043b\u0430\u044f)" if winner_val == "Radiant" else "Dire (\u0422\u044c\u043c\u0430)" if winner_val == "Dire" else ""
        summary_rows = [{
            "match_id": overview.get("match_id"),
            "duration": overview.get("duration", ""),
            "score": overview.get("score", ""),
            WINNER_COL: winner_label,
            "winner": winner_val,
            "players_parsed": overview.get("players_count", 0),
            "error": overview.get("error", ""),
        }]

        ws, headers, _ = _get_sheet_by_logical(wb, "summary", SUMMARY_COLS)
        if mid:
            deleted = _delete_match_rows(ws, headers, mid)
            if deleted and logger:
                logger.info(f"Match {mid}: removed {deleted} summary rows before append")
        _append_rows(ws, summary_rows, headers)

        players_rows = overview.get("players", []) or []
        ws, headers, _ = _get_sheet_by_logical(wb, "players", PLAYER_COLS)
        if mid:
            deleted = _delete_match_rows(ws, headers, mid)
            if deleted and logger:
                logger.info(f"Match {mid}: removed {deleted} player rows before append")
        _append_rows(ws, players_rows, headers)

    if update_lanes and lanes_rows is not None:
        ws, headers, _ = _get_sheet_by_logical(wb, "lanes", LANES_RAW_COLS)
        if mid:
            deleted = _delete_match_rows(ws, headers, mid)
            if deleted and logger:
                logger.info(f"Match {mid}: removed {deleted} lanes rows before append")
        _append_rows(ws, lanes_rows, headers)

        laning_stats = []
        for r in lanes_rows:
            laning_stats.append({
                "match_id": r.get("match_id"),
                "hero": r.get("hero"),
                "lane_outcome": r.get("lane_outcome"),
                "lane_team": r.get("lane_team"),
                "lane_detail": r.get("lane_detail"),
                "lane_simple": r.get("lane_simple"),
                "lane_side": r.get("lane_side"),
                "gpm_12": r.get("gpm_12"),
                "xpm_12": r.get("xpm_12"),
                "kills_12": r.get("k_12"),
                "deaths_12": r.get("d_12"),
                "assists_12": r.get("a_12"),
                "lh_4": r.get("lh_4"),
                "lh_8": r.get("lh_8"),
                "lh_12": r.get("lh_12"),
            })

        ws, headers, _ = _get_sheet_by_logical(wb, "laning", LANING_COLS)
        if mid:
            deleted = _delete_match_rows(ws, headers, mid)
            if deleted and logger:
                logger.info(f"Match {mid}: removed {deleted} laning rows before append")
        _append_rows(ws, laning_stats, headers)

        ws, headers, _ = _get_sheet_by_logical(wb, "laning_all", LANING_COLS)
        if mid:
            deleted = _delete_match_rows(ws, headers, mid)
            if deleted and logger:
                logger.info(f"Match {mid}: removed {deleted} full laning rows before append")
        _append_rows(ws, laning_stats, headers)

    wb.save(path)
    return True


def append_match_to_excel(overview, lanes_rows, output_path, logger=None):
    return update_match_in_excel(overview, lanes_rows, output_path, update_overview=True, update_lanes=True, logger=logger)


def save_to_excel(matches_overview, matches_lanes, output_path):
    """Save data to Excel:
    - Matches_Summary
    - Players_Overview
    - Lanes_Raw
    - Laning_Stats
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    sheet_summary = DEFAULT_SHEETS["summary"]
    sheet_players = DEFAULT_SHEETS["players"]
    sheet_lanes = DEFAULT_SHEETS["lanes"]
    sheet_laning = DEFAULT_SHEETS["laning"]
    sheet_laning_all = DEFAULT_SHEETS["laning_all"]

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        # --- summary ---
        summary_rows = []
        for m in matches_overview:
            winner_val = m.get("winner") or ""
            winner_label = "Radiant (\u0421\u0432\u0435\u0442\u043b\u0430\u044f)" if winner_val == "Radiant" else "Dire (\u0422\u044c\u043c\u0430)" if winner_val == "Dire" else ""
            summary_rows.append({
                "match_id": m.get("match_id"),
                "duration": m.get("duration", ""),
                "score": m.get("score", ""),
                WINNER_COL: winner_label,
                "winner": winner_val,
                "players_parsed": m.get("players_count", 0),
                "error": m.get("error", ""),
            })
        if summary_rows:
            df_summary = pd.DataFrame(summary_rows)
            for col in SUMMARY_COLS:
                if col not in df_summary.columns:
                    df_summary[col] = ""
            df_summary = df_summary[SUMMARY_COLS]
            df_summary.to_excel(w, sheet_name=sheet_summary, index=False)

        # --- players ---
        all_players = []
        for m in matches_overview:
            all_players.extend(m.get("players", []))
        if all_players:
            df_players = pd.DataFrame(all_players)
            for col in PLAYER_COLS:
                if col not in df_players.columns:
                    df_players[col] = ""
            df_players = df_players[PLAYER_COLS]
            df_players.to_excel(w, sheet_name=sheet_players, index=False)

        # --- lanes raw ---
        flat_lanes = [row for rows in matches_lanes for row in rows]
        if flat_lanes:
            df_lanes = pd.DataFrame(flat_lanes)
            for col in LANES_RAW_COLS:
                if col not in df_lanes.columns:
                    df_lanes[col] = ""
            df_lanes = df_lanes[LANES_RAW_COLS]
            df_lanes.to_excel(w, sheet_name=sheet_lanes, index=False)

        # --- laning stats ---
        laning_stats = []
        for rows in matches_lanes:
            for r in rows:
                laning_stats.append({
                    "match_id": r.get("match_id"),
                    "hero": r.get("hero"),
                    "lane_outcome": r.get("lane_outcome"),
                    "lane_team": r.get("lane_team"),
                    "lane_detail": r.get("lane_detail"),
                    "lane_simple": r.get("lane_simple"),
                    "lane_side": r.get("lane_side"),
                    "gpm_12": r.get("gpm_12"),
                    "xpm_12": r.get("xpm_12"),
                    "kills_12": r.get("k_12"),
                    "deaths_12": r.get("d_12"),
                    "assists_12": r.get("a_12"),
                    "lh_4": r.get("lh_4"),
                    "lh_8": r.get("lh_8"),
                    "lh_12": r.get("lh_12"),
                })
        df_laning = pd.DataFrame(laning_stats)
        if df_laning.empty:
            df_laning = pd.DataFrame(columns=LANING_COLS)
        else:
            for c in LANING_COLS:
                if c not in df_laning.columns:
                    df_laning[c] = ""
            df_laning = df_laning[LANING_COLS]

        df_laning.to_excel(w, sheet_name=sheet_laning, index=False)
        df_laning.to_excel(w, sheet_name=sheet_laning_all, index=False)

    print(f"Saved: {path.absolute()}")




def main():
    import os, argparse

    csv_path = r"c:\Users\79025\Downloads\Telegram Desktop\matches_ids_df.csv"
    output_excel = Path(__file__).resolve().parent / "dotabuff_matches_stats.xlsx"

    parser = argparse.ArgumentParser(description="Парсер матчей Dotabuff (PRO FIX)")
    parser.add_argument("--csv", default=os.environ.get("DOTABUFF_CSV", csv_path))
    parser.add_argument("--output", "-o", default=os.environ.get("DOTABUFF_OUTPUT"))
    parser.add_argument("--limit", "-n", type=int, default=0)
    parser.add_argument("--start", type=int, default=1, help="Стартовый номер матча (1-based)")
    parser.add_argument("--match-timeout", type=int, default=300, help="Макс. время на матч (сек)")
    parser.add_argument("--page-timeout-ms", type=int, default=120000, help="Макс. время на загрузку одной страницы (мс)")
    parser.add_argument("--page-retries", type=int, default=2, help="Кол-во повторов загрузки страницы при ошибках/таймаутах")
    parser.add_argument("--retry-wait", type=float, default=2.0, help="Пауза между повторами (сек)")
    parser.add_argument("--no-save-each", action="store_true", help="Не сохранять в Excel после каждого матча")
    parser.add_argument("--log", default="", help="Путь к лог-файлу")
    parser.add_argument("--headed", action="store_true")
    args = parser.parse_args()

    if args.output:
        output_excel = Path(args.output)

    save_each = not args.no_save_each
    log_path = args.log or str(Path(__file__).resolve().parent / "parser.log")
    logger = init_logging(log_path)

    existing_status = {}
    if output_excel.exists() and not save_each:
        logger.warning("Output Excel exists; forcing save_each to avoid data loss")
        save_each = True

    if save_each and output_excel.exists():
        try:
            deduplicate_excel(output_excel, logger=logger)
        except Exception as e:
            logger.warning(f"Excel dedup failed: {e}")
        try:
            existing_status = analyze_existing_excel(output_excel, logger=logger)
        except Exception as e:
            logger.warning(f"Excel analyze failed: {e}")

    match_ids = load_match_ids(args.csv)
    if args.limit and args.limit > 0:
        match_ids = match_ids[:args.limit]
        logger.info(f"Обрабатываем первые {args.limit} матчей")

    start_idx = max(1, int(args.start or 1)) - 1
    if start_idx >= len(match_ids):
        logger.error(f"Стартовый номер {args.start} превышает число матчей: {len(match_ids)}")
        return

    if start_idx > 0:
        match_ids = match_ids[start_idx:]
        logger.info(f"Стартуем с матча №{start_idx + 1} (ID {match_ids[0]})")

    logger.info(f"Матчей к обработке: {len(match_ids)}")

    if not HAS_PLAYWRIGHT:
        raise RuntimeError("Playwright не установлен. pip install playwright && playwright install chromium")

    debug_dir = Path(__file__).resolve().parent / "debug_html"
    debug_dir.mkdir(exist_ok=True)

    def fetch_with_retries(url, save_debug_path=None):
        last_err = None
        attempts = max(1, int(args.page_retries or 0)) + 1
        for attempt in range(1, attempts + 1):
            try:
                return fetch_dotabuff_with_playwright(
                    url,
                    save_debug_path=save_debug_path,
                    headed=args.headed,
                    max_total_ms=int(args.page_timeout_ms * attempt),
                    logger=logger
                )
            except Exception as e:
                last_err = e
                if attempt >= attempts:
                    break
                logger.warning(f"Fetch retry {attempt}/{attempts} for {url}: {e}")
                time.sleep(max(0.1, float(args.retry_wait)) * attempt)
        raise last_err

    matches_overview = []
    matches_lanes = []

    for i, mid in enumerate(match_ids):
        mid = str(mid).strip()
        if not mid:
            continue
        status = existing_status.get(mid) if save_each else None
        need_overview = True
        need_lanes = True
        if status:
            need_overview = status.get("need_overview", True)
            need_lanes = status.get("need_lanes", True)
            if not need_overview and not need_lanes:
                logger.info(f"[{i+1}/{len(match_ids)}] Match {mid} SKIP (already complete)")
                continue
            if not status.get("complete"):
                logger.info(f"[{i+1}/{len(match_ids)}] Match {mid} INCOMPLETE -> reparse")

        match_start = time.monotonic()
        def _check_match_timeout(stage):
            if args.match_timeout and args.match_timeout > 0:
                if time.monotonic() - match_start > args.match_timeout:
                    raise TimeoutError(f"Match {mid} timeout at stage: {stage}")
        logger.info(f"[{i+1}/{len(match_ids)}] Match {mid} START")
        url_overview = f"{BASE_URL}/matches/{mid}"
        url_lanes = f"{BASE_URL}/matches/{mid}/lanes"

        save_debug = str(debug_dir / f"match_{mid}_overview.html") if i == 0 else None
        overview = None
        lanes_rows = None

        if need_overview:
            try:
                logger.info(f"Match {mid}: fetch overview")
                _check_match_timeout("fetch_overview")
                html_overview = fetch_with_retries(url_overview, save_debug_path=save_debug)
                logger.info(f"Match {mid}: parse overview")
                _check_match_timeout("parse_overview")
                overview = parse_overview(html_overview, mid)
                matches_overview.append(overview)
            except Exception as e:
                logger.error(f"?????? ???????? overview {mid}: {e}")

            time.sleep(random.uniform(2.0, 4.0))

        if need_lanes:
            try:
                logger.info(f"Match {mid}: fetch lanes")
                _check_match_timeout("fetch_lanes")
                html_lanes = fetch_with_retries(url_lanes)
                logger.info(f"Match {mid}: parse lanes")
                _check_match_timeout("parse_lanes")
                lanes_rows = parse_lanes_tab(html_lanes, mid)
                matches_lanes.append(lanes_rows)

            except Exception as e:
                logger.error(f"?????? ???????? lanes {mid}: {e}")
                lanes_rows = []
                matches_lanes.append([])

        if save_each:
            try:
                logger.info(f"Match {mid}: update Excel")
                update_match_in_excel(
                    overview,
                    lanes_rows,
                    output_excel,
                    update_overview=bool(need_overview and overview is not None),
                    update_lanes=bool(need_lanes and lanes_rows is not None),
                    logger=logger
                )
            except Exception as e:
                logger.error(f"?????? ?????? Excel ??? {mid}: {e}")

        elapsed = time.monotonic() - match_start
        logger.info(f"[{i+1}/{len(match_ids)}] Match {mid} DONE in {elapsed:.1f}s")

        time.sleep(random.uniform(2.0, 4.0))

    if not save_each:
        save_to_excel(matches_overview, matches_lanes, output_excel)
    logger.info("Готово.")

if __name__ == "__main__":
    main()
